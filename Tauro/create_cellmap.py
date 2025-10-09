"""
Script para crear cellmap desde archivos Excel
Extrae todas las celdas con valores de todas las hojas
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def create_cellmap_from_excel(excel_path: str, output_path: str = None):
    """
    Crea un cellmap (diccionario de todas las celdas) desde un archivo Excel
    
    Args:
        excel_path: Ruta al archivo Excel (.xlsx, .xlsm)
        output_path: Ruta de salida para el JSON (opcional)
    
    Returns:
        Dict con estructura: {sheet_name: {cell_ref: value}}
    """
    print(f"📂 Abriendo archivo: {excel_path}")
    
    # Cargar workbook con data_only=True para obtener valores calculados
    wb = load_workbook(excel_path, data_only=True)
    
    cellmap = {}
    
    for sheet_name in wb.sheetnames:
        print(f"📄 Procesando hoja: {sheet_name}")
        sheet = wb[sheet_name]
        
        sheet_cells = {}
        cell_count = 0
        
        # Iterar por todas las filas y columnas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Crear referencia de celda (A1, B2, etc.)
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    sheet_cells[cell_ref] = cell.value
                    cell_count += 1
        
        cellmap[sheet_name] = sheet_cells
        print(f"   ✓ {cell_count} celdas extraídas")
    
    wb.close()
    
    # Determinar ruta de salida
    if output_path is None:
        input_path = Path(excel_path)
        output_dir = Path(__file__).parent / "output"
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / f"{input_path.stem}_cellmap.json"
    
    # Guardar cellmap
    print(f"\n💾 Guardando cellmap en: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(cellmap, f, ensure_ascii=False, indent=2, default=str)
    
    total_cells = sum(len(sheet) for sheet in cellmap.values())
    print(f"✅ Cellmap creado exitosamente")
    print(f"   📊 Total de hojas: {len(cellmap)}")
    print(f"   📊 Total de celdas: {total_cells}")
    
    return cellmap


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python create_cellmap.py <archivo.xlsx> [salida.json]")
        print("\nEjemplo:")
        print('  python create_cellmap.py "reporte.xlsx"')
        print('  python create_cellmap.py "reporte.xlsx" "output/reporte_cellmap.json"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    create_cellmap_from_excel(excel_file, output_file)
